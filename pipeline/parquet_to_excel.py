#!/usr/bin/env python3
"""
Convert paintings.parquet → paintings.xlsx with the same
styling as the original Excel output.
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

IN_PARQUET  = Path("paintings.parquet")
OUT_EXCEL   = Path("paintings.xlsx")

# ───────────────────── read & tidy ─────────────────────
df = pd.read_parquet(IN_PARQUET)

desired_cols = [
    "Title", "File Name", "Creator", "Year", "Material", "Dimensions",
    "Location", "Collection", "Movements", "Depicts",
    "Wikipedia URL", "Link Count",
    "Painting ID", "Creator ID", "Movement IDs",
]

# ensure any missing columns exist (empty)
for col in desired_cols:
    if col not in df.columns:
        df[col] = ""

df = df[desired_cols]
df["Link Count"] = pd.to_numeric(df["Link Count"], errors="coerce")
df.sort_values("Link Count", ascending=False, inplace=True)

# ───────────────────── write Excel ─────────────────────
with pd.ExcelWriter(OUT_EXCEL, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Paintings")
    ws = writer.sheets["Paintings"]

    # column widths
    for idx, col in enumerate(df.columns, 1):
        letter = get_column_letter(idx)
        max_len = max(len(col), df[col].astype(str).str.len().max())
        ws.column_dimensions[letter].width = min(max(max_len + 2, 8), 50)

        # numeric formats & hyperlink style
        if col == "Year":
            for cell in ws[letter][1:]:
                if cell.value and str(cell.value).isdigit():
                    cell.number_format = "0"
        elif col == "Link Count":
            for cell in ws[letter][1:]:
                if cell.value is not None:
                    cell.number_format = "#,##0"
        elif col in {"Wikipedia URL", "Painting ID", "Creator ID"}:
            for cell in ws[letter][1:]:
                if cell.value:
                    cell.style = "Hyperlink"

    # header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092",
                              end_color="366092",
                              fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center",
                                   vertical="center")

    ws.freeze_panes = "A2"

print(f"Excel file written → {OUT_EXCEL.resolve()}")

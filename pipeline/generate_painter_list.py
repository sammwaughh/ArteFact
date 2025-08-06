#!/usr/bin/env python3
"""
Build painters.xlsx from paintings.xlsx.

Input  (script 1 output)
------------------------
paintings.xlsx – sheet “Paintings”, column “Creator”

Output (for script 3)
---------------------
painters.xlsx – sheet “Painters” with a single column
• Artist  – original creator name (capitalisation preserved)

The sheet is auto-sized so script 3 can read it unchanged.
"""

from __future__ import annotations

from pathlib import Path
from typing import NoReturn

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ──────────────────────────── constants ───────────────────────────────────────
# work in the current directory (no dedicated Excel folder)
SOURCE_XLSX: Path = Path("paintings.xlsx")   # script-1 output
DEST_XLSX: Path = Path("painters.xlsx")      # new file


def autosize(ws: Worksheet) -> None:
    """Resize every column width to fit its longest cell."""
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value)) for cell in col if cell.value is not None)
        ws.column_dimensions[letter].width = max_len + 2


def copy_header_style(src_ws: Worksheet, dst_ws: Worksheet, col_idx: int = 1) -> None:
    """Copy width and header-cell style from src_ws to dst_ws for one column."""
    # copy column width
    letter = get_column_letter(col_idx)
    dst_ws.column_dimensions[letter].width = src_ws.column_dimensions[letter].width

    # copy cell style (font, alignment, fill, border)
    src_cell = src_ws.cell(row=1, column=col_idx)
    dst_cell = dst_ws.cell(row=1, column=col_idx)

    dst_cell.font = Font(**{k: v for k, v in vars(src_cell.font).items()
                            if not k.startswith('_')})
    dst_cell.alignment = Alignment(**{k: v for k, v in vars(src_cell.alignment).items()
                                      if not k.startswith('_')})
    dst_cell.fill = PatternFill(**{k: v for k, v in vars(src_cell.fill).items()
                                   if not k.startswith('_')})
    dst_cell.border = Border(**{k: v for k, v in vars(src_cell.border).items()
                                if not k.startswith('_')})


def main() -> NoReturn:
    """Extract creators from paintings.xlsx and save painters.xlsx."""
    # no directory to create – all files are in CWD

    source_path: Path = SOURCE_XLSX
    if not source_path.is_file():
        raise FileNotFoundError(f"{SOURCE_XLSX} (output of script 1) not found")

    # read the “Creator” column only
    df_paintings = pd.read_excel(
        source_path,
        sheet_name="Paintings",
        usecols=["Creator"],
    )
    creators = (
        df_paintings["Creator"].dropna().astype(str).str.strip().unique().tolist()
    )

    # ── NEW: discard rows that look like IRIs ---------------------------------
    creators = [
        c for c in creators if not c.lower().startswith(("http://", "https://"))
    ]

    # preserve original order of appearance
    artists = pd.Series(creators).drop_duplicates().tolist()

    data = [{"Artist": a} for a in artists]  # single column
    df_out = pd.DataFrame(data)

    with pd.ExcelWriter(DEST_XLSX, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Painters")

    # adjust / copy formatting
    src_wb = load_workbook(SOURCE_XLSX, read_only=False, data_only=True)
    src_ws = src_wb["Paintings"]

    dst_wb = load_workbook(DEST_XLSX)
    dst_ws = dst_wb["Painters"]

    copy_header_style(src_ws, dst_ws)      # keep same width & header style
    # optional: keep autosize for data rows if desired
    # autosize(dst_ws)

    dst_wb.save(DEST_XLSX)
    print(f"Created {DEST_XLSX} with {len(df_out)} unique artists.")


if __name__ == "__main__":
    main()

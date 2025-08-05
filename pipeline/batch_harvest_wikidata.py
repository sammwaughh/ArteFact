#!/usr/bin/env python3
"""
Harvest painting metadata from Wikidata and save to Excel, using
endpoint‑friendly settings that comply with WDQS rate limits.

Major changes versus previous revision
--------------------------------------
* Initial page size 250; sub‑chunk 100; only one worker touches WDQS.
* One‑second politeness delay after every HTTP request and five seconds
  between top‑level pages.
* Full support for HTTP 429: honour Retry‑After header when present.
* Slightly more informative User‑Agent string (contact + repo URL).
"""

from __future__ import annotations

import json
import logging
import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from itertools import islice
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import pandas as pd
import requests
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from tqdm import tqdm
from urllib3.util import Retry

# ─────────────────────────────────── constants ──────────────────────────────────
LIMIT: int = 32000
INITIAL_CHUNK_SIZE: int = 400  # ↑ 60% - fewer pages, still fast
MIN_CHUNK_SIZE: int = 100  # ↑ 100% - better fallback minimum
CHUNK_GROW_BACK_FACTOR: float = 1.2

SITELINKS_THRESHOLD: int = 1
SUBCHUNK_SIZE: int = 200  # ↑ 100% - fewer aggregation requests
WORKERS: int = 3  # ≤ 5 per WDQS policy  # ↑ 200% - optimal parallel sweet spot

# ───────────────────────────── contact constants ──────────────────────────────
CONTACT_EMAIL: str = "samjmwaugh@gmail.com"
ORG_URL: str = "https://github.com/sammwaughh/ArtContext"

PAUSE_BETWEEN_HTTP: float = (
    0.3  # polite delay after *every* HTTP  # ↓ 70% - still respectful
)
PAUSE_BETWEEN_PAGES: float = (
    1.5  # extra pause between main pages   # ↓ 70% - reasonable gap
)

MAX_ATTEMPTS: int = 6
REQUEST_TIMEOUT: Tuple[int, int] = (10, 75)

OUT_FILE: Path = Path("paintings.xlsx")

# ─────────────────────────────────── logging ────────────────────────────────────
os.makedirs("logs", exist_ok=True)
logging.basicConfig(
    filename="logs/wikidata_harvest.log",
    level=logging.INFO,
    filemode="w",
    format="%(asctime)s  %(levelname)s  %(message)s",
)
print("Harvesting painting metadata…")


# ──────────────────────────── helper: chunk an iterable ─────────────────────────
def chunked(it: Iterable, size: int):
    it = iter(it)
    while chunk := tuple(islice(it, size)):
        yield chunk


# ─────────────────────── requests session with retry policy ─────────────────────
def make_session() -> requests.Session:
    retry_cfg = Retry(
        total=5,
        backoff_factor=1.5,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"POST"}),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry_cfg)
    s = requests.Session()
    s.headers.update(
        {
            # WDQS maintainers want a *descriptive* UA with contact info
            "User-Agent": (
                f"ArtContextHarvester/0.2 ({ORG_URL}; mailto:{CONTACT_EMAIL})"
            ),
            "Accept": "application/sparql-results+json",
            "Content-Type": "application/x-www-form-urlencoded",
        }
    )
    s.mount("https://", adapter)
    return s


SESSION = make_session()
TRANSIENT = {502, 503, 504}


# ──────────────────────── robust SPARQL POST with retries ───────────────────────
def query_wd(query: str) -> dict:
    url = "https://query.wikidata.org/sparql"

    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            logging.info("HTTP POST to WDQS (attempt %d/%d)", attempt, MAX_ATTEMPTS)
            r = SESSION.post(url, data={"query": query}, timeout=REQUEST_TIMEOUT)
            if r.status_code == 429:  # rate‑limited
                retry_for = int(r.headers.get("Retry-After", "60"))
                logging.warning("429 Too Many Requests – waiting %s s", retry_for)
                time.sleep(retry_for)
                continue
            r.raise_for_status()

            # Safe JSON parsing with fallback
            try:
                return r.json()
            except json.JSONDecodeError as json_err:
                logging.error("JSON decode error: %s", json_err)
                logging.error("Response text (first 200 chars): %s", r.text[:200])
                # Treat JSON errors like network errors - retry with backoff
                if attempt < MAX_ATTEMPTS:
                    wait = 2**attempt
                    logging.warning("JSON parsing failed – retry in %s s", wait)
                    time.sleep(wait)
                    continue
                raise

        except requests.exceptions.HTTPError as exc:
            code = getattr(exc.response, "status_code", None)
            if code in TRANSIENT and attempt < MAX_ATTEMPTS:
                wait = 2**attempt
                logging.warning("%s from WDQS – retry in %s s", code, wait)
                time.sleep(wait)
                continue
            logging.error("Fatal HTTP error: %s", exc)
            raise

        except requests.exceptions.RequestException as exc:
            if attempt < MAX_ATTEMPTS:
                wait = 2**attempt
                logging.warning("Network error: %s – retry in %s s", exc, wait)
                time.sleep(wait)
                continue
            raise
        finally:
            # politeness delay: never hammer the server
            time.sleep(PAUSE_BETWEEN_HTTP)


# ────────────────────────────── basic‑page query ───────────────────────────────
def get_basic_records(offset: int, page_size: int, threshold: int):
    q = f"""
        SELECT ?painting ?paintingLabel ?creator ?creatorLabel
               ?inception ?wikipedia_url ?linkCount
               ?materialLabel ?height ?width ?locationLabel ?collectionLabel
        WHERE {{
          ?painting wdt:P31 wd:Q3305213 ;
                    wikibase:sitelinks ?linkCount .
          FILTER(?linkCount >= {threshold})
          OPTIONAL {{ ?painting wdt:P170 ?creator }}
          OPTIONAL {{ ?painting wdt:P571 ?inception }}
          OPTIONAL {{ ?painting wdt:P186 ?material }}
          OPTIONAL {{ ?painting wdt:P2048 ?height }}
          OPTIONAL {{ ?painting wdt:P2049 ?width }}
          OPTIONAL {{ ?painting wdt:P276 ?location }}
          OPTIONAL {{ ?painting wdt:P195 ?collection }}
          OPTIONAL {{
            ?paintingArticle schema:about ?painting ;
                             schema:inLanguage "en" ;
                             schema:isPartOf <https://en.wikipedia.org/> .
            BIND(?paintingArticle AS ?wikipedia_url)
          }}
          SERVICE wikibase:label {{ bd:serviceParam wikibase:language "en". }}
        }}
        LIMIT {page_size} OFFSET {offset}
        """
    try:
        data = query_wd(q)
    except (
        requests.exceptions.ReadTimeout,
        requests.exceptions.HTTPError,
        json.JSONDecodeError,
    ) as exc:
        # halve page size on time‑out, 504, or JSON parsing errors
        if page_size > MIN_CHUNK_SIZE:
            new_size = max(page_size // 2, MIN_CHUNK_SIZE)
            logging.warning("Query failed (%s) – retrying with size %d", exc, new_size)
            return get_basic_records(offset, new_size, threshold)
        raise

    bindings = data.get("results", {}).get("bindings", [])
    records, pids = {}, []
    for b in bindings:
        pid = b["painting"]["value"]
        records[pid] = {
            "Painting ID": pid,
            "Title": b.get("paintingLabel", {}).get("value", ""),
            "Creator ID": b.get("creator", {}).get("value", ""),
            "Creator": b.get("creatorLabel", {}).get("value", ""),
            "Inception": b.get("inception", {}).get("value", ""),
            "Wikipedia URL": b.get("wikipedia_url", {}).get("value", ""),
            "Link Count": int(b["linkCount"]["value"]),
            # Add the missing fields
            "Material": b.get("materialLabel", {}).get("value", ""),
            "Height": b.get("height", {}).get("value", ""),
            "Width": b.get("width", {}).get("value", ""),
            "Location": b.get("locationLabel", {}).get("value", ""),
            "Collection": b.get("collectionLabel", {}).get("value", ""),
        }
        pids.append(pid)
    return records, pids


# ───────────────────────── aggregate fields in sub‑chunks ──────────────────────
def get_agg_fields(painting_ids: List[str]) -> Dict[str, dict]:
    agg: Dict[str, dict] = {}

    def one_chunk(ids: Tuple[str, ...]) -> Dict[str, dict]:
        vals = " ".join(f"<{x}>" for x in ids)
        q = f"""
            SELECT ?painting
                   (GROUP_CONCAT(DISTINCT ?depictsLabel;  separator=", ") AS ?depicts)
                   (GROUP_CONCAT(DISTINCT ?movementLabel; separator=", ") AS ?movs)
                   (GROUP_CONCAT(DISTINCT ?movement;      separator=", ") AS ?movIDs)
            WHERE {{
              VALUES ?painting {{ {vals} }}
              OPTIONAL {{
                  ?painting wdt:P180 ?depicts .
                  ?depicts  rdfs:label ?depictsLabel .
                  FILTER(LANG(?depictsLabel) = "en")
              }}
              OPTIONAL {{
                  ?painting wdt:P135 ?movement .
                  ?movement rdfs:label ?movementLabel .
                  FILTER(LANG(?movementLabel) = "en")
              }}
            }}
            GROUP BY ?painting
            """
        try:
            d = query_wd(q)
        except requests.exceptions.HTTPError as exc:
            if exc.response and exc.response.status_code in TRANSIENT and len(ids) > 1:
                mid = len(ids) // 2
                return {**one_chunk(ids[:mid]), **one_chunk(ids[mid:])}
            raise
        res = {}
        for b in d.get("results", {}).get("bindings", []):
            pid = b["painting"]["value"]
            res[pid] = {
                "Depicts": b.get("depicts", {}).get("value", ""),
                "Movements": b.get("movs", {}).get("value", ""),
                "Movement IDs": b.get("movIDs", {}).get("value", ""),
            }
        return res

    # single worker keeps us under 5‑parallel limit
    with ThreadPoolExecutor(max_workers=WORKERS) as pool:
        futs = [
            pool.submit(one_chunk, tuple(ch))
            for ch in chunked(painting_ids, SUBCHUNK_SIZE)
        ]
        for f in as_completed(futs):
            agg.update(f.result())
    return agg


# ─────────────────────────────────── main loop ─────────────────────────────────
all_rows: List[dict] = []
offset, page_size = 0, INITIAL_CHUNK_SIZE
bar = tqdm(total=LIMIT, unit=" rows")

try:
    while len(all_rows) < LIMIT:
        logging.info("Page offset %d, size %d", offset, page_size)
        basic, ids = get_basic_records(offset, page_size, SITELINKS_THRESHOLD)
        if not basic:
            break

        agg = get_agg_fields(ids)
        merged = [
            {
                **basic[pid],
                **agg.get(pid, {"Depicts": "", "Movements": "", "Movement IDs": ""}),
            }
            for pid in basic
        ]

        existing = {r["Painting ID"] for r in all_rows}
        new = [r for r in merged if r["Painting ID"] not in existing]
        all_rows.extend(new)
        bar.update(len(new))

        offset += len(basic)
        # gentle ramp‑up once stable
        if page_size < INITIAL_CHUNK_SIZE and len(new) == len(basic):
            page_size = min(int(page_size * CHUNK_GROW_BACK_FACTOR), INITIAL_CHUNK_SIZE)

        time.sleep(PAUSE_BETWEEN_PAGES)

except KeyboardInterrupt:
    logging.warning("Interrupted – writing partial results.")

finally:
    bar.close()

logging.info("Total unique paintings: %d", len(all_rows))

# ───────────────────────────── dataframe & excel ──────────────────────────────
df = pd.DataFrame(all_rows)
df["File Name"] = (
    df["Painting ID"].str.extract(r"([^/]+)$", expand=False).fillna("") + "_0.png"
)
df["Year"] = pd.to_numeric(df["Inception"].str[:4], errors="coerce")

# Create combined dimensions field
df["Dimensions"] = ""
if "Height" in df.columns and "Width" in df.columns:
    mask = (
        df["Height"].notna()
        & df["Width"].notna()
        & (df["Height"] != "")
        & (df["Width"] != "")
    )
    if mask.any():
        df.loc[mask, "Dimensions"] = (
            df.loc[mask, "Height"].astype(str)
            + " × "
            + df.loc[mask, "Width"].astype(str)
            + " cm"
        )

desired_cols = [
    "Title",
    "File Name",
    "Creator",
    "Year",
    "Material",
    "Dimensions",  # Combine height x width
    "Location",
    "Collection",
    "Movements",
    "Depicts",
    "Wikipedia URL",
    "Link Count",
    "Painting ID",
    "Creator ID",
    "Movement IDs",
]

# Ensure all desired columns exist with empty string defaults
for col in desired_cols:
    if col not in df.columns:
        df[col] = ""

# Select only the desired columns in the specified order
df = df[desired_cols]
df["Link Count"] = pd.to_numeric(df["Link Count"], errors="coerce")
df.sort_values("Link Count", ascending=False, inplace=True)

with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as w:
    df.to_excel(w, index=False, sheet_name="Paintings")
    ws = w.sheets["Paintings"]

    # Enhanced column formatting
    for i, col in enumerate(df.columns, 1):
        col_letter = get_column_letter(i)

        # Calculate optimal width based on content
        max_length = max(
            df[col].astype(str).str.len().max() if not df[col].empty else 0, len(col)
        )
        # Set width with reasonable bounds
        width = min(max(max_length + 2, 8), 50)
        ws.column_dimensions[col_letter].width = width

        # Special formatting for specific columns
        if col == "Year":
            for cell in ws[col_letter][1:]:  # Skip header
                if cell.value and str(cell.value).isdigit():
                    cell.number_format = "0"
        elif col == "Link Count":
            for cell in ws[col_letter][1:]:  # Skip header
                if cell.value:
                    cell.number_format = "#,##0"
        elif col in ["Wikipedia URL", "Painting ID", "Creator ID"]:
            for cell in ws[col_letter][1:]:  # Skip header
                if cell.value:
                    cell.style = "Hyperlink"

    # Style the header row
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze the header row
    ws.freeze_panes = "A2"

print("Done – saved", OUT_FILE)

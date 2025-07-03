import os
import re
import json
import time
import threading
import psutil
import requests
import pandas as pd
import concurrent.futures
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

my_email = "xlct43@durham.ac.uk"

# --- Utility: Setup logger ---
def setup_logger(name, log_file, level=logging.INFO, fmt="%(asctime)s - %(levelname)s - %(message)s"):
    """Set up a logger that writes to log_file."""
    logger = logging.getLogger(name)
    logger.setLevel(level)
    # Remove any existing handlers (if re‑using the same name)
    if logger.hasHandlers():
        logger.handlers.clear()
    fh = logging.FileHandler(log_file)
    formatter = logging.Formatter(fmt)
    fh.setFormatter(formatter)
    logger.addHandler(fh)
    return logger

# Global logger variables (will be re‑assigned for each painter)
fetch_logger = None
excel_logger = None
download_logger = None
cpu_logger = None

# Global list to store download time data (reset for each painter)
download_time_data = []

# --- Utility functions ---
def convert_to_str(value):
    """Convert dictionaries or lists to a JSON string; otherwise, return the value."""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value

def safe_json_parse(cell_value):
    """If the cell value is already a dict or list, return it; otherwise, try to parse it as JSON."""
    if isinstance(cell_value, (dict, list)):
        return cell_value
    try:
        return json.loads(cell_value)
    except Exception:
        return None

def sanitize_filename(name):
    """Remove characters not allowed in file names."""
    return re.sub(r'[\\/*?:"<>|]', "", name).strip()

# --- Core functions ---
def fetch_works(painter_name):
    """
    Fetch all works from OpenAlex that meet the following criteria:
      - language is English (language:en)
      - the work is open access (is_oa:true)
      - the work has one of the specified topic IDs in its topics list
      - the work's searchable metadata mentions the painter_name
    """
    base_url = "https://api.openalex.org/works"
    topics_ids = (
        "T14092|T14191|T12372|T14469|T12680|T14366|T13922|T12444|"
        "T13133|T12179|T13342|T12632|T14002|T14322"
    )
    filter_str = f"language:en,is_oa:true,topics.id:{topics_ids}"
    cursor = "*"
    per_page = 200
    works = []
    page_count = 0
    headers = {"User-Agent": f"MyPythonClient (mailto:{my_email})"}
    
    fetch_logger.info(f"Starting query for works mentioning '{painter_name}'...")
    while cursor:
        page_count += 1
        fetch_logger.info(f"Querying page {page_count}...")
        params = {
            "filter": filter_str,
            "search": painter_name,
            "per_page": per_page,
            "cursor": cursor,
            "mailto": my_email
        }
        while True:
            response = requests.get(base_url, params=params, headers=headers)
            if response.status_code == 429:
                fetch_logger.warning("Rate limit exceeded. Sleeping for 60 seconds...")
                time.sleep(60)
            else:
                break
        if response.status_code != 200:
            fetch_logger.error(f"Error: Received status code {response.status_code}")
            fetch_logger.error(f"Response content: {response.text}")
            break
        data = response.json()
        results = data.get("results", [])
        works.extend(results)
        fetch_logger.info(f"Page {page_count}: Retrieved {len(results)} works.")
        cursor = data.get("meta", {}).get("next_cursor")
        if not cursor:
            fetch_logger.info("No further pages found.")
            break
        time.sleep(1)
    fetch_logger.info(f"Query complete. Total pages: {page_count}. Total works retrieved: {len(works)}")
    return works

def get_candidate_links(row):
    """Given a row (from the Filtered DataFrame), extract candidate download links."""
    candidates = []
    best_oa = safe_json_parse(row.get('best_oa_location'))
    if best_oa and isinstance(best_oa, dict):
        if best_oa.get('pdf_url'):
            candidates.append(best_oa['pdf_url'])
        elif best_oa.get('landing_page_url'):
            candidates.append(best_oa['landing_page_url'])
    oa = safe_json_parse(row.get('open_access'))
    if oa and isinstance(oa, dict) and oa.get('oa_url'):
        candidates.append(oa['oa_url'])
    primary = safe_json_parse(row.get('primary_location'))
    if primary and isinstance(primary, dict):
        if primary.get('pdf_url'):
            candidates.append(primary['pdf_url'])
        elif primary.get('landing_page_url'):
            candidates.append(primary['landing_page_url'])
    locs = safe_json_parse(row.get('locations'))
    if locs and isinstance(locs, list):
        for loc in locs:
            if isinstance(loc, dict):
                if loc.get('pdf_url'):
                    candidates.append(loc['pdf_url'])
                elif loc.get('landing_page_url'):
                    candidates.append(loc['landing_page_url'])
    return list(dict.fromkeys(candidates))

def get_best_and_backup_links(row):
    """
    Returns a tuple (best_link, backup_link) for a given row from the Filtered DataFrame.
    Strategy: choose the first candidate containing '.pdf' (case‑insensitive) as best_link,
    and the next candidate (if any) as backup_link.
    """
    candidates = get_candidate_links(row)
    best_link = ""
    backup_link = ""
    for link in candidates:
        if '.pdf' in link.lower():
            best_link = link
            break
    if not best_link and candidates:
        best_link = candidates[0]
    for link in candidates:
        if link != best_link:
            backup_link = link
            break
    return best_link, backup_link

def create_excel_file(painter_name, works):
    """
    Using the fetched works, create three DataFrames:
      - df_main: all columns
      - df_filtered: a subset of columns
      - df_downloadable: with columns Title, Relevance Score, Best Link, Back Up Link, OpenAlexID, Type
    Write all three DataFrames to separate sheets in a single Excel file (saved in ExcelFiles directory).
    """
    df_main = pd.DataFrame(works)
    
    filtered_columns = [
        'title', 'relevance_score', 'id', 'doi', 
        'primary_location', 'type', 'open_access', 
        'locations', 'best_oa_location'
    ]
    df_filtered = df_main[filtered_columns].copy()
    
    downloadable_data = []
    for _, row in df_filtered.iterrows():
        title = row.get('title', '')
        relevance = row.get('relevance_score', 0)
        openalex_id = row.get('id', '')
        work_type = row.get('type', '')
        best_link, backup_link = get_best_and_backup_links(row)
        downloadable_data.append({
            'Title': title,
            'Relevance Score': relevance,
            'Best Link': best_link,
            'Back Up Link': backup_link,
            'OpenAlexID': openalex_id,
            'Type': work_type
        })
    df_downloadable = pd.DataFrame(downloadable_data)
    
    filename = os.path.join("ExcelFiles", f"{painter_name.lower()}_works.xlsx")
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df_main.to_excel(writer, index=False, sheet_name="Main")
        df_filtered.to_excel(writer, index=False, sheet_name="Filtered")
        df_downloadable.to_excel(writer, index=False, sheet_name="Downloadable")
    
    wb = load_workbook(filename)
    ws_main = wb["Main"]
    for col in ws_main.columns:
        col_letter = get_column_letter(col[0].column)
        max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
        ws_main.column_dimensions[col_letter].width = max_length + 2
    for sheet in ["Filtered", "Downloadable"]:
        ws = wb[sheet]
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = 35
    wb.save(filename)
    excel_logger.info(f"Excel file '{filename}' created with sheets: Main, Filtered, and Downloadable.")
    return filename

def download_pdf(index, title, best_link, backup_link, openalex_id, work_type, directory):
    """
    Attempts to download a PDF using best_link first; if that fails, tries backup_link.
    Measures and logs the time taken (in seconds, rounded to three decimals) for a successful download.
    Saves the file as "{index}-{sanitized_title}.pdf" in the directory.
    Records the PDF name in download_time_data.
    """
    sanitized_title = sanitize_filename(title)
    pdf_filename = f"{index}-{sanitized_title}.pdf"
    filepath = os.path.join(directory, pdf_filename)
    start_time = time.perf_counter()
    for link in [best_link, backup_link]:
        if link and isinstance(link, str) and link.strip():
            try:
                download_logger.info(f"Row {index}: Attempting download from: {link}")
                response = requests.get(link, stream=True, timeout=20)
                if response.status_code == 200:
                    with open(filepath, 'wb') as f:
                        for chunk in response.iter_content(chunk_size=8192):
                            f.write(chunk)
                    end_time = time.perf_counter()
                    elapsed_sec = round(end_time - start_time, 3)
                    download_logger.info(f"Row {index}: Successfully downloaded as: {filepath} in {elapsed_sec} seconds")
                    download_time_data.append({
                        "PDF Name": pdf_filename,
                        "Time": elapsed_sec,
                        "Type": work_type
                    })
                    return True
                else:
                    download_logger.warning(f"Row {index}: Failed with status {response.status_code} on link: {link}")
            except Exception as e:
                download_logger.error(f"Row {index}: Error downloading from {link}: {e}")
    download_logger.error(f"Row {index}: All download attempts failed (ID: {openalex_id}).")
    return False

def monitor_cpu(stop_event, interval=1):
    """
    Monitors and logs CPU usage every `interval` seconds until stop_event is set.
    """
    cpu_usages = []
    while not stop_event.is_set():
        usage = psutil.cpu_percent(interval=interval)
        cpu_usages.append(usage)
        cpu_logger.info(f"Current CPU usage: {usage}%")
    avg = sum(cpu_usages)/len(cpu_usages) if cpu_usages else 0
    cpu_logger.info(f"Average CPU usage during download: {avg:.2f}%")

def append_download_time_sheet(excel_file, painter_name, total_runtime):
    """
    Appends a new sheet "Download Times" to the existing Excel file.
    The sheet contains a table (PDF Name, Time, Type) sorted by Time descending,
    with a final row showing the total program runtime.
    """
    if download_time_data:
        df_time = pd.DataFrame(download_time_data, columns=["PDF Name", "Time", "Type"])
        df_time = df_time.sort_values(by="Time", ascending=False).reset_index(drop=True)
        total_row = pd.DataFrame({
            "PDF Name": ["Total Program Runtime"],
            "Time": [f"{total_runtime:.3f} seconds"],
            "Type": [""]
        })
        df_time = pd.concat([df_time, total_row], ignore_index=True)
        
        wb = load_workbook(excel_file)
        if "Download Times" in wb.sheetnames:
            ws_del = wb["Download Times"]
            wb.remove(ws_del)
        ws = wb.create_sheet(title="Download Times")
        
        headers = list(df_time.columns)
        ws.append(headers)
        for row in df_time.itertuples(index=False, name=None):
            ws.append(list(row))
        
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
            ws.column_dimensions[col_letter].width = max_length + 2
        wb.save(excel_file)
        download_logger.info(f"Download time data appended as a new sheet in '{excel_file}'.")
        print(f"Download time data appended as a new sheet in '{excel_file}'.")
    else:
        download_logger.info("No download time data to append.")

def download_all_pdfs(excel_file, painter_name, max_workers):
    """
    Reads the "Downloadable" sheet, creates the painter’s PDF directory, and downloads PDFs in parallel.
    Only downloads works with a relevance score > 1.
    Also launches a thread to monitor CPU usage.
    After downloads complete, appends a "Download Times" sheet to the Excel file.
    """
    df = pd.read_excel(excel_file, sheet_name="Downloadable", engine="openpyxl")
    pdf_dir = os.path.join("PDFs", painter_name.lower())
    os.makedirs(pdf_dir, exist_ok=True)
    
    stop_event = threading.Event()
    cpu_thread = threading.Thread(target=monitor_cpu, args=(stop_event,), daemon=True)
    cpu_thread.start()
    
    start_time = time.time()
    download_logger.info("Starting parallel PDF download tasks...")
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = []
        for i, row in enumerate(df.itertuples(index=False)):
            relevance = row[1]
            if relevance <= 1:
                download_logger.info(f"Row {i}: Skipping due to low relevance score ({relevance}).")
                continue
            title = row[0]
            best_link = row[2]
            backup_link = row[3]
            openalex_id = row[4]
            work_type = row[5]
            download_logger.info(f"Submitting download task for row {i}: {title}")
            futures.append(
                executor.submit(download_pdf, i, title, best_link, backup_link, openalex_id, work_type, pdf_dir)
            )
        download_logger.info("Waiting for all download tasks to complete...")
        concurrent.futures.wait(futures)
    stop_event.set()
    cpu_thread.join()
    end_time = time.time()
    total_runtime = end_time - start_time
    download_logger.info(f"All downloads attempted in {total_runtime:.2f} seconds.")
    print(f"All downloads attempted in {total_runtime:.2f} seconds.")
    
    # Append the download time sheet to the Excel file.
    append_download_time_sheet(excel_file, painter_name, total_runtime)

# --- Painter processing ---
def process_painter(painter, max_workers):
    """
    For one painter, set up individual loggers (with file names using the painter name and timestamp),
    then fetch works, create the Excel file, download PDFs, and log the overall runtime.
    """
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    global fetch_logger, excel_logger, download_logger, cpu_logger, download_time_data
    # Set up individual log files (stored in respective directories)
    fetch_logger = setup_logger("fetch", os.path.join("fetch-logs", f"{painter}-{timestamp}.log"))
    excel_logger = setup_logger("excel", os.path.join("excel-logs", f"{painter}-{timestamp}.log"))
    download_logger = setup_logger("download", os.path.join("download-logs", f"{painter}-{timestamp}.log"))
    cpu_logger = setup_logger("cpu", os.path.join("cpu-logs", f"{painter}-{timestamp}.log"))
    
    # Reset download time data for this painter
    download_time_data = []
    
    overall_start = time.time()
    print(f"Processing painter: {painter}")
    fetch_logger.info(f"Processing painter: {painter}")
    
    works = fetch_works(painter)
    print(f"Fetched {len(works)} works for {painter}")
    
    excel_file = create_excel_file(painter, works)
    print(f"Excel file created: {excel_file}")
    
    download_all_pdfs(excel_file, painter, max_workers)
    
    overall_end = time.time()
    runtime = overall_end - overall_start
    download_logger.info(f"Total program runtime for {painter}: {runtime:.2f} seconds.")
    print(f"Total runtime for {painter}: {runtime:.2f} seconds.")

# --- Main execution ---
def main():
    # Create required directories if they don't exist.
    for dirname in ["ExcelFiles", "PDFs", "excel-logs", "cpu-logs", "download-logs", "fetch-logs"]:
        os.makedirs(dirname, exist_ok=True)
    
    # List of painters to process (replace with your 100 names as desired)
    painters = [
        "davinci", "da vinci"
    ]
    
    # Configure max_workers (you can adjust this value)
    max_workers = 10
    
    # Process each painter one after the other.
    for painter in painters:
        process_painter(painter, max_workers)

if __name__ == "__main__":
    main()
